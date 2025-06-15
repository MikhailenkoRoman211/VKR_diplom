from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Optional

import numpy as np
import pandas as pd
from pypfopt import EfficientFrontier, risk_models

@dataclass
class Optimizer:

    X: pd.DataFrame
    y_pred: np.ndarray
    price_df: pd.DataFrame
    lookback: int = 252
    min_avg_volume: int | None = None
    top_n: int = 10

    def _latest_per_ticker(self) -> pd.DataFrame:
        return self.X.groupby("SECID_code").tail(1)

    def _code2secid(self) -> Dict[int, str]:
        cat = self.price_df["SECID"].astype("category")
        return {code: secid for code, secid in enumerate(cat.cat.categories)}

    def _recent_ret_matrix(self) -> pd.DataFrame:
        pdf = self.price_df.copy()
        pdf["TRADEDATE"] = pd.to_datetime(pdf["TRADEDATE"])
        pdf.sort_values(["SECID", "TRADEDATE"], inplace=True)

        cutoff = pdf["TRADEDATE"].max() - pd.Timedelta(days=int(self.lookback * 1.5))
        pdf = pdf.loc[pdf["TRADEDATE"] >= cutoff]

        pdf["ret"] = pdf.groupby("SECID")["CLOSE"].pct_change(fill_method=None)
        return pdf.pivot(index="TRADEDATE", columns="SECID", values="ret").tail(self.lookback)

    def compute(self) -> pd.Series:
        X_last = self._latest_per_ticker()
        code2secid = self._code2secid()

        y_series = pd.Series(self.y_pred, index=self.X.index)
        mu = y_series.loc[X_last.index].copy()
        secids_pred = [code2secid[int(c)] for c in X_last["SECID_code"]]
        mu.index = secids_pred

        R = self._recent_ret_matrix()

        if self.min_avg_volume:
            liquid = (
                self.price_df.groupby("SECID")["VOLUME"]
                .mean()
                .loc[lambda s: s >= self.min_avg_volume]
                .index
            )
            R = R[liquid]

        common = list(set(mu.index) & set(R.columns))
        mu, R = mu[common], R[common]

        Sigma = risk_models.sample_cov(R)
        ef = EfficientFrontier(mu, Sigma, weight_bounds=(0, 0.15))
        _ = ef.max_sharpe()
        w_all = pd.Series(ef.clean_weights()).sort_values(ascending=False)

        weights = w_all.head(self.top_n)
        weights = weights / weights.sum()

        self._mu_series = mu
        self._Sigma_df  = Sigma
        self._ef = ef
        return weights

    def save_report(self, weights: pd.Series, path: str | Path) -> None:
        import numpy as np
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        secids = weights.index
        mu    = self._mu_series.loc[secids]
        cov_mat = self._Sigma_df.loc[secids, secids]
        sigma   = pd.Series(np.sqrt(np.diag(cov_mat)), index=secids)
        avg_vol = (
            self.price_df[self.price_df["SECID"].isin(secids)]
            .groupby("SECID")["VOLUME"].mean()
            .reindex(secids)
        )

        df_w = pd.DataFrame({
            "доля": weights.values,
            "прогноз, %": mu.values,
            "риск, %": sigma.values,
            "сред. объём": avg_vol.values,
        }, index=secids).round(6)

        port_ret, port_risk, sharpe = self._ef.portfolio_performance()
        df_perf = pd.DataFrame(
            {"metric": ["expected_return", "volatility", "sharpe"],
             "value":  [port_ret,          port_risk,    sharpe]}
        )

        risks, rets = [], []
        if hasattr(self._ef, "efficient_risk"):
            for s in np.linspace(0.01, 0.4, 50):
                try:
                    self._ef.efficient_risk(s)
                    r, v, _ = self._ef.portfolio_performance()
                    risks.append(v); rets.append(r)
                except Exception:
                    pass
        df_front = pd.DataFrame({"risk": risks, "return": rets})

        wb = Workbook()
        ws_w = wb.active; ws_w.title = "Weights"
        for r in dataframe_to_rows(df_w, index=True, header=True):
            ws_w.append(r)

        ws_p = wb.create_sheet("Performance")
        for r in dataframe_to_rows(df_perf, index=False, header=True):
            ws_p.append(r)

        if not df_front.empty:
            ws_f = wb.create_sheet("Frontier")
            for r in dataframe_to_rows(df_front, index=False, header=True):
                ws_f.append(r)

        wb.save(path)
        print(f"[Optimizer] report saved → {path}")
